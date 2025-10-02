from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Count, Sum
from django.db import transaction
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
import django_filters
import csv
import openpyxl
from datetime import datetime
from io import BytesIO

from .models import (
    TipoAjuste, CuentaContable, AjusteFinanciero,
    HistorialAjuste, ArchivoAdjunto, ComentarioAjuste
)
from .serializers import (
    TipoAjusteSerializer, CuentaContableSerializer,
    AjusteFinancieroListSerializer, AjusteFinancieroDetailSerializer,
    AjusteFinancieroCreateUpdateSerializer, CambiarEstadoAjusteSerializer,
    HistorialAjusteSerializer, ArchivoAdjuntoSerializer,
    ComentarioAjusteSerializer, ExportarAjustesSerializer
)

class AjusteFinancieroFilter(django_filters.FilterSet):
    """Filtros para AjusteFinanciero"""
    fecha_inicio = django_filters.DateFilter(field_name="fecha_ajuste", lookup_expr='gte')
    fecha_fin = django_filters.DateFilter(field_name="fecha_ajuste", lookup_expr='lte')
    monto_min = django_filters.NumberFilter(field_name="monto", lookup_expr='gte')
    monto_max = django_filters.NumberFilter(field_name="monto", lookup_expr='lte')
    usuario_creador = django_filters.CharFilter(field_name="usuario_creador__username")
    tipo_ajuste = django_filters.ModelMultipleChoiceFilter(queryset=TipoAjuste.objects.all())
    estado = django_filters.MultipleChoiceFilter(choices=AjusteFinanciero.ESTADO_CHOICES)
    
    class Meta:
        model = AjusteFinanciero
        fields = [
            'estado', 'prioridad', 'tipo_ajuste', 'cuenta_debito', 'cuenta_credito',
            'moneda', 'usuario_creador', 'centro_costo'
        ]

class TipoAjusteViewSet(viewsets.ModelViewSet):
    """ViewSet para TipoAjuste"""
    queryset = TipoAjuste.objects.all()
    serializer_class = TipoAjusteSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['nombre', 'descripcion']
    ordering_fields = ['nombre', 'created_at']
    ordering = ['nombre']
    filterset_fields = ['activo']

class CuentaContableViewSet(viewsets.ModelViewSet):
    """ViewSet para CuentaContable"""
    queryset = CuentaContable.objects.all()
    serializer_class = CuentaContableSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['codigo', 'nombre', 'descripcion']
    ordering_fields = ['codigo', 'nombre', 'created_at']
    ordering = ['codigo']
    filterset_fields = ['tipo_cuenta', 'activo']

class AjusteFinancieroViewSet(viewsets.ModelViewSet):
    """ViewSet principal para AjusteFinanciero"""
    queryset = AjusteFinanciero.objects.select_related(
        'tipo_ajuste', 'cuenta_debito', 'cuenta_credito',
        'usuario_creador', 'usuario_aprobador', 'usuario_procesador'
    ).prefetch_related('historial', 'archivos', 'comentarios')
    
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_class = AjusteFinancieroFilter
    search_fields = [
        'numero_ajuste', 'concepto', 'descripcion', 'justificacion',
        'numero_documento_origen', 'referencia_externa'
    ]
    ordering_fields = [
        'numero_ajuste', 'fecha_ajuste', 'fecha_valor', 'monto',
        'estado', 'prioridad', 'created_at'
    ]
    ordering = ['-fecha_ajuste', '-numero_ajuste']
    
    def get_serializer_class(self):
        """Retorna el serializer apropiado según la acción"""
        if self.action == 'list':
            return AjusteFinancieroListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return AjusteFinancieroCreateUpdateSerializer
        else:
            return AjusteFinancieroDetailSerializer
    
    def get_queryset(self):
        """Personalizar queryset según permisos del usuario"""
        queryset = super().get_queryset()
        user = self.request.user
        
        # Los administradores ven todos los ajustes
        if user.is_superuser:
            return queryset
        
        # Los usuarios normales solo ven los ajustes que crearon o donde están involucrados
        return queryset.filter(
            Q(usuario_creador=user) |
            Q(usuario_aprobador=user) |
            Q(usuario_procesador=user)
        )
    
    @action(detail=True, methods=['post'])
    def cambiar_estado(self, request, pk=None):
        """Cambiar el estado de un ajuste"""
        ajuste = self.get_object()
        serializer = CambiarEstadoAjusteSerializer(
            data=request.data,
            context={'ajuste': ajuste, 'request': request}
        )
        
        if serializer.is_valid():
            nuevo_estado = serializer.validated_data['nuevo_estado']
            comentario = serializer.validated_data.get('comentario', '')
            
            with transaction.atomic():
                # Guardar historial
                HistorialAjuste.objects.create(
                    ajuste=ajuste,
                    estado_anterior=ajuste.estado,
                    estado_nuevo=nuevo_estado,
                    usuario=request.user,
                    comentario=comentario
                )
                
                # Actualizar estado y fechas
                estado_anterior = ajuste.estado
                ajuste.estado = nuevo_estado
                
                if nuevo_estado == 'APROBADO' and estado_anterior != 'APROBADO':
                    ajuste.usuario_aprobador = request.user
                    ajuste.fecha_aprobacion = timezone.now()
                elif nuevo_estado == 'PROCESADO' and estado_anterior != 'PROCESADO':
                    ajuste.usuario_procesador = request.user
                    ajuste.fecha_procesamiento = timezone.now()
                
                ajuste.save()
            
            return Response({
                'message': f'Estado cambiado a {nuevo_estado} exitosamente',
                'nuevo_estado': nuevo_estado
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def subir_archivo(self, request, pk=None):
        """Subir archivo adjunto a un ajuste"""
        ajuste = self.get_object()
        
        # Verificar que el ajuste pueda ser editado
        if not ajuste.puede_ser_editado:
            return Response(
                {'error': 'No se pueden subir archivos a este ajuste en su estado actual'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = ArchivoAdjuntoSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save(ajuste=ajuste)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def agregar_comentario(self, request, pk=None):
        """Agregar comentario a un ajuste"""
        ajuste = self.get_object()
        serializer = ComentarioAjusteSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save(ajuste=ajuste)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtener historial de cambios de un ajuste"""
        ajuste = self.get_object()
        historial = ajuste.historial.all()
        serializer = HistorialAjusteSerializer(historial, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def mis_ajustes(self, request):
        """Obtener ajustes creados por el usuario actual"""
        ajustes = self.get_queryset().filter(usuario_creador=request.user)
        
        # Aplicar filtros y paginación
        filtered_queryset = self.filter_queryset(ajustes)
        page = self.paginate_queryset(filtered_queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(filtered_queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def pendientes_aprobacion(self, request):
        """Obtener ajustes pendientes de aprobación"""
        # Solo usuarios con permisos pueden ver esto
        if not request.user.has_perm('adjustments.can_approve'):
            return Response(
                {'error': 'No tiene permisos para ver ajustes pendientes'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        ajustes = self.get_queryset().filter(estado='PENDIENTE')
        
        # Aplicar filtros y paginación
        filtered_queryset = self.filter_queryset(ajustes)
        page = self.paginate_queryset(filtered_queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(filtered_queryset, many=True)
        return Response(serializer.data)

class ExportAjustesView(APIView):
    """Vista para exportar ajustes en diferentes formatos"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        serializer = ExportarAjustesSerializer(data=request.data)
        
        if serializer.is_valid():
            formato = serializer.validated_data['formato']
            fecha_inicio = serializer.validated_data.get('fecha_inicio')
            fecha_fin = serializer.validated_data.get('fecha_fin')
            estados = serializer.validated_data.get('estado', [])
            tipos_ajuste = serializer.validated_data.get('tipo_ajuste', [])
            
            # Construir queryset
            queryset = AjusteFinanciero.objects.select_related(
                'tipo_ajuste', 'cuenta_debito', 'cuenta_credito', 'usuario_creador'
            )
            
            # Aplicar filtros
            if fecha_inicio:
                queryset = queryset.filter(fecha_ajuste__gte=fecha_inicio)
            if fecha_fin:
                queryset = queryset.filter(fecha_ajuste__lte=fecha_fin)
            if estados:
                queryset = queryset.filter(estado__in=estados)
            if tipos_ajuste:
                queryset = queryset.filter(tipo_ajuste__in=tipos_ajuste)
            
            # Generar archivo según formato
            if formato == 'csv':
                return self._export_csv(queryset)
            elif formato == 'excel':
                return self._export_excel(queryset)
            elif formato == 'pdf':
                return self._export_pdf(queryset)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def _export_csv(self, queryset):
        """Exportar a CSV"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="ajustes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Encabezados
        writer.writerow([
            'Número Ajuste', 'Fecha Ajuste', 'Fecha Valor', 'Tipo Ajuste',
            'Cuenta Débito', 'Cuenta Crédito', 'Monto', 'Moneda',
            'Concepto', 'Estado', 'Prioridad', 'Usuario Creador'
        ])
        
        # Datos
        for ajuste in queryset:
            writer.writerow([
                ajuste.numero_ajuste,
                ajuste.fecha_ajuste.strftime('%Y-%m-%d %H:%M:%S'),
                ajuste.fecha_valor.strftime('%Y-%m-%d'),
                ajuste.tipo_ajuste.get_nombre_display(),
                f"{ajuste.cuenta_debito.codigo} - {ajuste.cuenta_debito.nombre}",
                f"{ajuste.cuenta_credito.codigo} - {ajuste.cuenta_credito.nombre}",
                str(ajuste.monto),
                ajuste.moneda,
                ajuste.concepto,
                ajuste.get_estado_display(),
                ajuste.get_prioridad_display(),
                ajuste.usuario_creador.get_full_name() or ajuste.usuario_creador.username
            ])
        
        return response
    
    def _export_excel(self, queryset):
        """Exportar a Excel"""
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Ajustes Financieros"
        
        # Encabezados
        headers = [
            'Número Ajuste', 'Fecha Ajuste', 'Fecha Valor', 'Tipo Ajuste',
            'Cuenta Débito', 'Cuenta Crédito', 'Monto', 'Moneda',
            'Concepto', 'Estado', 'Prioridad', 'Usuario Creador'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Datos
        for row, ajuste in enumerate(queryset, 2):
            worksheet.cell(row=row, column=1, value=ajuste.numero_ajuste)
            worksheet.cell(row=row, column=2, value=ajuste.fecha_ajuste)
            worksheet.cell(row=row, column=3, value=ajuste.fecha_valor)
            worksheet.cell(row=row, column=4, value=ajuste.tipo_ajuste.get_nombre_display())
            worksheet.cell(row=row, column=5, value=f"{ajuste.cuenta_debito.codigo} - {ajuste.cuenta_debito.nombre}")
            worksheet.cell(row=row, column=6, value=f"{ajuste.cuenta_credito.codigo} - {ajuste.cuenta_credito.nombre}")
            worksheet.cell(row=row, column=7, value=float(ajuste.monto))
            worksheet.cell(row=row, column=8, value=ajuste.moneda)
            worksheet.cell(row=row, column=9, value=ajuste.concepto)
            worksheet.cell(row=row, column=10, value=ajuste.get_estado_display())
            worksheet.cell(row=row, column=11, value=ajuste.get_prioridad_display())
            worksheet.cell(row=row, column=12, value=ajuste.usuario_creador.get_full_name() or ajuste.usuario_creador.username)
        
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ajustes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
    
    def _export_pdf(self, queryset):
        """Exportar a PDF - implementación básica"""
        # Aquí podrías usar librerías como reportlab o weasyprint
        # Por ahora retornamos un error indicando que no está implementado
        return Response(
            {'error': 'Exportación a PDF no implementada aún'},
            status=status.HTTP_501_NOT_IMPLEMENTED
        )

class ImportAjustesView(APIView):
    """Vista para importar ajustes desde archivo"""
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        archivo = request.FILES.get('archivo')
        
        if not archivo:
            return Response(
                {'error': 'Debe proporcionar un archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar tipo de archivo
        if not archivo.name.endswith(('.csv', '.xlsx')):
            return Response(
                {'error': 'Solo se permiten archivos CSV o Excel'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            if archivo.name.endswith('.csv'):
                resultados = self._import_csv(archivo, request.user)
            else:
                resultados = self._import_excel(archivo, request.user)
            
            return Response(resultados)
        
        except Exception as e:
            return Response(
                {'error': f'Error al procesar archivo: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _import_csv(self, archivo, usuario):
        """Importar desde CSV"""
        # Implementación básica - se puede expandir según necesidades
        return {'message': 'Importación CSV no implementada aún'}
    
    def _import_excel(self, archivo, usuario):
        """Importar desde Excel"""
        # Implementación básica - se puede expandir según necesidades
        return {'message': 'Importación Excel no implementada aún'}

class BulkDeleteView(APIView):
    """Vista para eliminación en lote"""
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        ajuste_ids = request.data.get('ajuste_ids', [])
        
        if not ajuste_ids:
            return Response(
                {'error': 'Debe proporcionar al menos un ID de ajuste'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que el usuario pueda eliminar estos ajustes
        ajustes = AjusteFinanciero.objects.filter(id__in=ajuste_ids)
        
        # Verificar permisos
        for ajuste in ajustes:
            if not ajuste.puede_ser_editado:
                return Response(
                    {'error': f'El ajuste {ajuste.numero_ajuste} no puede ser eliminado en su estado actual'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Eliminar ajustes
        eliminados = ajustes.count()
        ajustes.delete()
        
        return Response({
            'message': f'Se eliminaron {eliminados} ajustes exitosamente',
            'eliminados': eliminados
        })
